import json
import re
import zipfile
from io import BytesIO
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.core.config import OUTPUTS_DIR
from app.schemas.design import ConfigFieldSpec, ConfigTableSpec, DesignData


def create_design_output_id() -> str:
    now = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"config_tables_{now}"


def get_design_output_dir() -> Path:
    output_dir = OUTPUTS_DIR / "design"
    output_dir.mkdir(parents=True, exist_ok=True)
    return output_dir


def save_design_outputs(design_data: DesignData) -> dict:
    output_id = create_design_output_id()
    output_dir = get_design_output_dir()

    json_filename = f"{output_id}.json"
    excel_filename = f"{output_id}.xlsx"
    excel_zip_filename = f"{output_id}_excel.zip"
    godot_zip_filename = f"{output_id}_godot.zip"

    json_path = output_dir / json_filename
    excel_path = output_dir / excel_filename
    excel_zip_path = output_dir / excel_zip_filename
    godot_zip_path = output_dir / godot_zip_filename

    save_design_json(design_data, json_path)
    save_design_excel(design_data, excel_path)
    save_excel_package(design_data, excel_path, json_path, excel_zip_path, output_id)
    save_godot_package(design_data, godot_zip_path, output_id)

    return {
        "output_id": output_id,
        "json_path": f"outputs/design/{json_filename}",
        "excel_path": f"outputs/design/{excel_filename}",
        "excel_zip_path": f"outputs/design/{excel_zip_filename}",
        "godot_zip_path": f"outputs/design/{godot_zip_filename}",
    }


def save_design_json(design_data: DesignData, json_path: Path) -> None:
    payload = {
        "title": design_data.title,
        "gameplay_summary": design_data.gameplay_summary,
        "tables": [
            {
                "name": table.name,
                "display_name": table.display_name,
                "purpose": table.purpose,
                "engine_usage": table.engine_usage,
                "fields": [field.model_dump() for field in table.fields],
                "rows": table.rows,
                "notes": table.notes,
            }
            for table in design_data.tables
        ],
        "export_notes": design_data.export_notes,
    }

    with json_path.open("w", encoding="utf-8") as file:
        json.dump(payload, file, ensure_ascii=False, indent=2)


def save_design_excel(design_data: DesignData, excel_path: Path) -> None:
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    add_overview_sheet(workbook, design_data)

    for table in design_data.tables:
        add_config_table_sheet(workbook, table)

    workbook.save(excel_path)


def add_overview_sheet(workbook: Workbook, design_data: DesignData) -> None:
    sheet = workbook.create_sheet("Overview")
    sheet.append(["Field", "Value"])
    sheet.append(["Title", design_data.title])
    sheet.append(["Gameplay Summary", design_data.gameplay_summary])
    sheet.append(["Tables", ", ".join(table.name for table in design_data.tables)])
    sheet.append(["Export Notes", "\n".join(design_data.export_notes)])

    style_sheet(sheet)
    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 100
    for cell in sheet["B"]:
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def add_config_table_sheet(workbook: Workbook, table: ConfigTableSpec) -> None:
    sheet_name = safe_sheet_name(table.name)
    sheet = workbook.create_sheet(sheet_name)
    add_config_table_rows(sheet, table)


def add_config_table_rows(sheet, table: ConfigTableSpec) -> None:
    field_names = [field.name for field in table.fields]
    sheet.append(field_names)
    sheet.append([field.type for field in table.fields])
    sheet.append([field.description for field in table.fields])

    for row in table.rows:
        sheet.append([format_excel_value(row.get(field_name, "")) for field_name in field_names])

    style_sheet(sheet)
    type_fill = PatternFill("solid", fgColor="D9EAF7")
    desc_fill = PatternFill("solid", fgColor="EAF4DD")

    for cell in sheet[2]:
        cell.fill = type_fill
        cell.font = Font(color="1F2937", bold=True)
    for cell in sheet[3]:
        cell.fill = desc_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    sheet.freeze_panes = "A4"
    for column_cells in sheet.columns:
        column_letter = column_cells[0].column_letter
        sheet.column_dimensions[column_letter].width = 24
    for row_cells in sheet.iter_rows():
        for cell in row_cells:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def save_excel_package(
    design_data: DesignData,
    excel_path: Path,
    json_path: Path,
    zip_path: Path,
    output_id: str,
) -> None:
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as package:
        package.write(excel_path, arcname=f"{output_id}.xlsx")
        package.write(json_path, arcname=f"{output_id}.json")
        package.writestr("README.md", render_excel_package_readme(design_data))

        for table in design_data.tables:
            table_payload = {
                "name": table.name,
                "display_name": table.display_name,
                "purpose": table.purpose,
                "engine_usage": table.engine_usage,
                "fields": [field.model_dump() for field in table.fields],
                "rows": table.rows,
                "notes": table.notes,
            }
            package.writestr(
                f"tables/{safe_file_name(table.name)}.json",
                json.dumps(table_payload, ensure_ascii=False, indent=2),
            )
            package.writestr(
                f"tables/{safe_file_name(table.name)}.xlsx",
                render_single_table_excel(table),
            )


def render_excel_package_readme(design_data: DesignData) -> str:
    lines = [
        f"# {design_data.title}",
        "",
        design_data.gameplay_summary,
        "",
        "## Package Contents",
        "",
        "- One `.xlsx` workbook with all generated config tables.",
        "- One full `.json` package for runtime import or tooling.",
        "- Per-table `.json` and `.xlsx` files under `tables/`.",
        "",
        "## Tables",
        "",
    ]

    for table in design_data.tables:
        lines.append(f"- {table.name}: {table.purpose}")

    lines.append("")
    return "\n".join(lines)


def render_single_table_excel(table: ConfigTableSpec) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = safe_sheet_name(table.name)
    add_config_table_rows(sheet, table)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def save_godot_package(design_data: DesignData, zip_path: Path, output_id: str) -> None:
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as package:
        package.writestr("config_table_resource.gd", render_godot_resource_script())
        package.writestr("README.md", render_godot_readme(design_data))

        for table in design_data.tables:
            file_name = f"tables/{safe_file_name(table.name)}.tres"
            package.writestr(file_name, render_table_tres(table))

        package.writestr(
            f"{output_id}.json",
            json.dumps(design_data.model_dump(), ensure_ascii=False, indent=2),
        )


def render_godot_resource_script() -> str:
    return """extends Resource
class_name ConfigTableResource

@export var table_name: String = ""
@export_multiline var purpose: String = ""
@export var fields: Array[Dictionary] = []
@export var rows: Array[Dictionary] = []
"""


def render_godot_readme(design_data: DesignData) -> str:
    lines = [
        f"# {design_data.title}",
        "",
        design_data.gameplay_summary,
        "",
        "## Usage",
        "",
        "1. Copy `config_table_resource.gd` into your Godot project.",
        "2. Copy the `.tres` files from `tables/` into a data folder.",
        "3. Load a table with `load(\"res://path/to/Items.tres\")`.",
        "4. Read `rows` as dictionaries in your gameplay systems.",
        "",
        "## Tables",
        "",
    ]

    for table in design_data.tables:
        lines.append(f"- {table.name}: {table.purpose}")

    lines.append("")
    return "\n".join(lines)


def render_table_tres(table: ConfigTableSpec) -> str:
    return "\n".join(
        [
            '[gd_resource type="Resource" script_class="ConfigTableResource" load_steps=2 format=3]',
            "",
            '[ext_resource type="Script" path="res://config_table_resource.gd" id="1"]',
            "",
            "[resource]",
            'script = ExtResource("1")',
            f'table_name = "{escape_godot_string(table.name)}"',
            f'purpose = "{escape_godot_string(table.purpose)}"',
            f"fields = {to_godot_variant_array([field.model_dump() for field in table.fields])}",
            f"rows = {to_godot_variant_array(table.rows)}",
            "",
        ]
    )


def to_godot_variant_array(value: list[dict[str, Any]]) -> str:
    return json.dumps(value, ensure_ascii=False)


def escape_godot_string(value: str) -> str:
    return value.replace("\\", "\\\\").replace('"', '\\"').replace("\n", "\\n")


def format_excel_value(value: Any) -> Any:
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return value


def style_sheet(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def safe_sheet_name(name: str) -> str:
    cleaned = re.sub(r"[\[\]\:\*\?\/\\]", "_", name).strip() or "Table"
    return cleaned[:31]


def safe_file_name(name: str) -> str:
    cleaned = re.sub(r"[^a-zA-Z0-9_\-]+", "_", name).strip("_")
    return cleaned or "table"
