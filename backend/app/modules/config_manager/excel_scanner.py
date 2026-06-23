from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

from app.schemas.configs import (
    ConfigIssue,
    ConfigScanResponse,
    ConfigScanSummary,
    ConfigSheetSummary,
    ConfigWorkbookItem,
)


SKIPPED_DIRS = {
    ".git",
    ".godot",
    ".venv",
    "__pycache__",
    "build",
    "dist",
    "library",
    "node_modules",
    "temp",
}

def scan_config_folder(folder_path: str) -> ConfigScanResponse:
    root = Path(folder_path).expanduser().resolve()
    if not root.exists():
        raise ValueError("Folder does not exist.")
    if not root.is_dir():
        raise ValueError("Path must be a folder.")

    workbooks: list[ConfigWorkbookItem] = []
    issues: list[ConfigIssue] = []
    skipped_temp_files = 0

    for path in _iter_excel_files(root):
        if path.name.startswith("~$"):
            skipped_temp_files += 1
            continue

        workbook_item = _scan_workbook(root, path)
        workbooks.append(workbook_item)
        issues.extend(workbook_item.issues)
        for sheet in workbook_item.sheets:
            issues.extend(sheet.issues)

    return ConfigScanResponse(
        root_path=str(root),
        summary=ConfigScanSummary(
            workbook_count=len(workbooks),
            sheet_count=sum(workbook.sheet_count for workbook in workbooks),
            issue_count=len(issues),
            skipped_temp_files=skipped_temp_files,
        ),
        workbooks=workbooks,
        issues=issues,
    )


def _iter_excel_files(root: Path):
    stack = [root]
    while stack:
        current = stack.pop()
        try:
            children = list(current.iterdir())
        except OSError:
            continue

        for child in children:
            if child.is_dir():
                if child.name.lower() not in SKIPPED_DIRS:
                    stack.append(child)
                continue

            if child.suffix.lower() == ".xlsx":
                yield child


def _scan_workbook(root: Path, path: Path) -> ConfigWorkbookItem:
    relative_path = str(path.relative_to(root))
    workbook_issues: list[ConfigIssue] = []
    sheets: list[ConfigSheetSummary] = []

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        workbook_issues.append(
            ConfigIssue(
                severity="error",
                code="workbook_load_failed",
                message=f"Unable to read workbook: {exc}",
                workbook=relative_path,
            )
        )
        return ConfigWorkbookItem(
            name=path.name,
            path=str(path),
            relative_path=relative_path,
            sheet_count=0,
            sheets=[],
            issues=workbook_issues,
        )

    try:
        for sheet in workbook.worksheets:
            sheets.append(_scan_sheet(relative_path, sheet))
    finally:
        workbook.close()

    return ConfigWorkbookItem(
        name=path.name,
        path=str(path),
        relative_path=relative_path,
        sheet_count=len(sheets),
        sheets=sheets,
        issues=workbook_issues,
    )


def _scan_sheet(workbook_name: str, sheet) -> ConfigSheetSummary:
    row_count = int(sheet.max_row or 0)
    column_count = int(sheet.max_column or 0)
    headers = _read_headers(sheet, column_count)

    return ConfigSheetSummary(
        name=sheet.title,
        row_count=row_count,
        column_count=column_count,
        headers=headers,
        issues=_diagnose_sheet(workbook_name, sheet.title, row_count, column_count, headers),
    )


def _read_headers(sheet, column_count: int) -> list[str]:
    if column_count <= 0 or not sheet.max_row:
        return []

    first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    return ["" if value is None else str(value).strip() for value in first_row[:column_count]]


def _diagnose_sheet(
    workbook_name: str,
    sheet_name: str,
    row_count: int,
    column_count: int,
    headers: list[str],
) -> list[ConfigIssue]:
    issues: list[ConfigIssue] = []
    location = {"workbook": workbook_name, "sheet": sheet_name}

    if row_count == 0 or column_count == 0 or not any(headers):
        issues.append(
            ConfigIssue(
                severity="warning",
                code="empty_sheet",
                message="Sheet has no usable header row.",
                **location,
            )
        )
        return issues

    normalized = [header.strip().lower() for header in headers if header.strip()]
    if "id" not in normalized:
        issues.append(
            ConfigIssue(
                severity="warning",
                code="missing_id",
                message="Sheet is missing an id column.",
                **location,
            )
        )

    blank_count = len([header for header in headers if not header.strip()])
    if blank_count:
        issues.append(
            ConfigIssue(
                severity="warning",
                code="blank_header",
                message=f"Header row contains {blank_count} blank column(s).",
                **location,
            )
        )

    duplicates = sorted([name for name, count in Counter(normalized).items() if count > 1])
    for duplicate in duplicates:
        issues.append(
            ConfigIssue(
                severity="error",
                code="duplicate_header",
                message=f"Duplicate header: {duplicate}",
                column=duplicate,
                **location,
            )
        )

    return issues
